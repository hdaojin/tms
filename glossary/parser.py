from __future__ import annotations

from collections import defaultdict
from typing import BinaryIO

from openpyxl import load_workbook

from core.constants import GLOSSARY_WORKBOOK_MAX_ROWS

from .normalization import english_comparison_key, normalize_display_text


TERM_HEADER = "term"
TRANSLATION_HEADER = "translation (used for smartcat - mandatory)"
ACRONYM_HEADER = "acronym"
IGNORED_HEADERS = {"序号", "no.", "no", "number"}


class GlossaryWorkbookError(Exception):
    def __init__(self, errors: list[str] | str):
        self.errors = [errors] if isinstance(errors, str) else errors
        super().__init__("; ".join(self.errors))


def _header_key(value: object) -> str:
    return normalize_display_text(value).casefold()


def _row_payload(sheet_name: str, row_number: int, english: str, acronym: str, translation: str) -> dict:
    return {
        "sheet": sheet_name,
        "row_number": row_number,
        "english_term": english,
        "english_key": english_comparison_key(english),
        "acronym": acronym,
        "chinese_translation": translation,
    }


def parse_smartcat_workbook(file_obj: BinaryIO) -> dict:
    try:
        file_obj.seek(0)
        workbook = load_workbook(file_obj, read_only=True, data_only=False)
    except Exception as exc:
        raise GlossaryWorkbookError(f"无法读取 XLSX 文件：{exc}") from exc

    matched_sheets: list[tuple[object, dict[str, int], list[str]]] = []
    empty_template_sheets: list[str] = []
    for worksheet in workbook.worksheets:
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1), ())
        headers = [_header_key(cell.value) for cell in first_row]
        mapping = {header: index for index, header in enumerate(headers) if header}
        if TERM_HEADER in mapping and TRANSLATION_HEADER in mapping:
            has_data = any(
                any(normalize_display_text(cell.value) for cell in row)
                for row in worksheet.iter_rows(min_row=2)
            )
            if has_data:
                matched_sheets.append((worksheet, mapping, headers))
            else:
                empty_template_sheets.append(worksheet.title)

    if not matched_sheets:
        workbook.close()
        if empty_template_sheets:
            raise GlossaryWorkbookError("Smartcat 数据工作表中没有可导入的数据行。")
        raise GlossaryWorkbookError(
            "未找到 Smartcat 数据工作表；第 1 行必须包含 Term 和 "
            "Translation (used for Smartcat - mandatory)。"
        )
    if len(matched_sheets) > 1:
        names = "、".join(worksheet.title for worksheet, _mapping, _headers in matched_sheets)
        workbook.close()
        raise GlossaryWorkbookError(f"发现多个 Smartcat 数据工作表（{names}），请只保留一个。")

    worksheet, mapping, headers = matched_sheets[0]
    warnings: list[str] = []
    if empty_template_sheets:
        warnings.append(f"已跳过空模板工作表：{'、'.join(empty_template_sheets)}。")
    recognized = {TERM_HEADER, TRANSLATION_HEADER, ACRONYM_HEADER} | IGNORED_HEADERS
    unknown_columns = [str(index + 1) for index, header in enumerate(headers) if header and header not in recognized]
    if unknown_columns:
        warnings.append(f"工作表 {worksheet.title} 的未识别列 {', '.join(unknown_columns)} 已忽略。")

    errors: list[str] = []
    candidates: list[dict] = []
    blank_rows = 0
    data_rows = 0
    unheaded_data_columns: set[int] = set()
    acronym_index = mapping.get(ACRONYM_HEADER)
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        if any(normalize_display_text(cell.value) for cell in row):
            data_rows += 1
            if data_rows > GLOSSARY_WORKBOOK_MAX_ROWS:
                errors.append(f"数据行不能超过 {GLOSSARY_WORKBOOK_MAX_ROWS} 条。")
                break
        term_cell = row[mapping[TERM_HEADER]] if mapping[TERM_HEADER] < len(row) else None
        translation_cell = row[mapping[TRANSLATION_HEADER]] if mapping[TRANSLATION_HEADER] < len(row) else None
        acronym_cell = row[acronym_index] if acronym_index is not None and acronym_index < len(row) else None
        for index, cell in enumerate(row):
            if (index >= len(headers) or not headers[index]) and normalize_display_text(cell.value):
                unheaded_data_columns.add(index + 1)
        if any(cell.data_type == "f" for cell in row):
            errors.append(f"{worksheet.title} 第 {row_number} 行：词条字段不允许使用公式。")
            continue

        english = normalize_display_text(term_cell.value if term_cell is not None else "")
        translation = normalize_display_text(translation_cell.value if translation_cell is not None else "")
        acronym = normalize_display_text(acronym_cell.value if acronym_cell is not None else "")
        if not english and not translation and not acronym:
            blank_rows += 1
            continue
        if not english:
            errors.append(f"{worksheet.title} 第 {row_number} 行：Term 不能为空。")
        if not translation:
            errors.append(f"{worksheet.title} 第 {row_number} 行：Translation 不能为空。")
        if len(english) > 255:
            errors.append(f"{worksheet.title} 第 {row_number} 行：Term 不能超过 255 个字符。")
        if len(acronym) > 100:
            errors.append(f"{worksheet.title} 第 {row_number} 行：Acronym 不能超过 100 个字符。")
        if len(translation) > 2000:
            errors.append(f"{worksheet.title} 第 {row_number} 行：Translation 不能超过 2000 个字符。")
        if english and translation and len(english) <= 255 and len(acronym) <= 100 and len(translation) <= 2000:
            candidates.append(_row_payload(worksheet.title, row_number, english, acronym, translation))

    workbook.close()
    if unheaded_data_columns:
        warnings.append(
            f"工作表 {worksheet.title} 的无表头数据列 "
            f"{', '.join(str(value) for value in sorted(unheaded_data_columns))} 已忽略。"
        )
    if errors:
        raise GlossaryWorkbookError(errors)

    grouped: dict[str, list[dict]] = defaultdict(list)
    for candidate in candidates:
        grouped[candidate["english_key"]].append(candidate)

    groups: list[dict] = []
    identical_duplicates = 0
    conflicting_duplicates = 0
    for key, rows in grouped.items():
        unique_rows: list[dict] = []
        seen_values: set[tuple[str, str, str]] = set()
        for row in rows:
            signature = (
                row["english_key"],
                english_comparison_key(row["acronym"]),
                row["chinese_translation"],
            )
            if signature in seen_values:
                identical_duplicates += 1
                continue
            seen_values.add(signature)
            unique_rows.append(row)
        if len(unique_rows) > 1:
            conflicting_duplicates += 1
        groups.append({"english_key": key, "options": unique_rows})

    groups.sort(key=lambda item: item["english_key"])
    if blank_rows:
        warnings.append(f"已忽略 {blank_rows} 个空白数据行。")
    if identical_duplicates:
        warnings.append(f"已合并 {identical_duplicates} 个内容完全相同的重复行。")
    return {
        "sheet_name": worksheet.title,
        "groups": groups,
        "warnings": warnings,
        "counts": {
            "source_rows": len(candidates),
            "unique_terms": len(groups),
            "identical_duplicates": identical_duplicates,
            "conflicting_duplicates": conflicting_duplicates,
        },
    }
