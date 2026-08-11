import tempfile
from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook

from standards.models import SkillProject

from glossary.models import GlossaryEntry, GlossaryEntryProposal, GlossaryImport, ProfessionalGlossary
from glossary.parser import GlossaryWorkbookError, parse_smartcat_workbook
from glossary.services import confirm_glossary_import, create_glossary_import


TERM = "Term"
TRANSLATION = "Translation (used for Smartcat - mandatory)"


def workbook_upload(rows, *, headers=None, extra_sheet=False, name="terms.xlsx"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers or ["序号", TERM, "Acronym", TRANSLATION, None])
    for row in rows:
        worksheet.append(row)
    if extra_sheet:
        duplicate = workbook.create_sheet("Sheet2")
        duplicate.append([TERM, TRANSLATION])
        duplicate.append(["Switch", "交换机"])
    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        name,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class SmartcatWorkbookParserTests(TestCase):
    def test_empty_smartcat_template_sheet_is_ignored(self):
        workbook = Workbook()
        template = workbook.active
        template.title = "空模板"
        template.append([TERM, TRANSLATION])
        data = workbook.create_sheet("数据")
        data.append([TERM, TRANSLATION])
        data.append(["Router", "路由器"])
        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile("terms.xlsx", output.getvalue())

        payload = parse_smartcat_workbook(upload)

        self.assertEqual(payload["sheet_name"], "数据")
        self.assertTrue(any("空模板" in warning for warning in payload["warnings"]))

    def test_cleans_rows_merges_identical_duplicates_and_warns_about_ignored_data(self):
        upload = workbook_upload(
            [
                [1, " Router  ", "", "路由器", "不用保留"],
                [2, "router", "", "路由器", "ignored"],
                [3, "Switch\u00a0  Core", "SW", "核心交换机", None],
            ]
        )

        payload = parse_smartcat_workbook(upload)

        self.assertEqual(payload["counts"]["unique_terms"], 2)
        self.assertEqual(payload["counts"]["identical_duplicates"], 1)
        self.assertEqual(payload["groups"][1]["options"][0]["english_term"], "Switch Core")
        self.assertTrue(any("无表头数据列 5" in warning for warning in payload["warnings"]))

    def test_conflicting_duplicates_are_kept_for_explicit_selection(self):
        payload = parse_smartcat_workbook(
            workbook_upload([[1, "Router", "", "路由器"], [2, "router", "", "路由设备"]])
        )

        self.assertEqual(payload["counts"]["conflicting_duplicates"], 1)
        self.assertEqual(len(payload["groups"][0]["options"]), 2)

    def test_formula_or_missing_required_value_blocks_the_batch(self):
        with self.assertRaises(GlossaryWorkbookError) as context:
            parse_smartcat_workbook(
                workbook_upload([[1, "=UPPER(\"router\")", "", "路由器"], [2, "Switch", "", ""]])
            )

        self.assertEqual(len(context.exception.errors), 2)
        self.assertTrue(any("公式" in error for error in context.exception.errors))
        self.assertTrue(any("Translation 不能为空" in error for error in context.exception.errors))

    def test_multiple_matching_sheets_are_rejected(self):
        with self.assertRaisesMessage(GlossaryWorkbookError, "多个 Smartcat"):
            parse_smartcat_workbook(workbook_upload([[1, "Router", "", "路由器"]], extra_sheet=True))


class GlossaryImportServiceTests(TestCase):
    def setUp(self):
        self.private_dir = tempfile.TemporaryDirectory()
        self.settings_override = override_settings(PRIVATE_MEDIA_ROOT=self.private_dir.name)
        self.settings_override.enable()
        self.addCleanup(self.private_dir.cleanup)
        self.addCleanup(self.settings_override.disable)
        self.user = get_user_model().objects.create_user(username="manager", password="test")
        project = SkillProject.objects.create(code="39", name="信息网络布线")
        self.glossary = ProfessionalGlossary.objects.create(
            skill_project=project,
            name="WSC 2026",
            created_by=self.user,
        )

    def test_invalid_batch_is_privately_archived_and_cannot_be_confirmed(self):
        record = create_glossary_import(
            self.glossary,
            workbook_upload([[1, "Router", "", ""]]),
            user=self.user,
        )

        self.assertEqual(record.status, GlossaryImport.Status.INVALID)
        self.assertTrue(record.source_file.storage.exists(record.source_file.name))
        self.assertTrue(record.parsed_payload["errors"])
        with self.assertRaisesMessage(ValidationError, "已确认或已过期"):
            confirm_glossary_import(record, {}, user=self.user)

    def test_confirm_supports_selection_overwrite_pending_skip_and_audit(self):
        existing = GlossaryEntry.objects.create(
            glossary=self.glossary,
            english_term="Router",
            chinese_translation="旧释义",
            is_active=False,
        )
        GlossaryEntryProposal.objects.create(
            glossary=self.glossary,
            english_term="Switch",
            chinese_translation="待审交换机",
            submitted_by=self.user,
        )
        record = create_glossary_import(
            self.glossary,
            workbook_upload(
                [
                    [1, "router", "RTR", "新释义"],
                    [2, "Cable", "", "电缆"],
                    [3, "cable", "", "线缆"],
                    [4, "Switch", "", "交换机"],
                ]
            ),
            user=self.user,
        )
        indexes = {group["english_key"]: index for index, group in enumerate(record.parsed_payload["groups"])}
        decisions = {
            "overwrite_all": False,
            "overwrite": [str(indexes["router"])],
            "choices": {str(indexes["cable"]): 1},
        }

        confirmed = confirm_glossary_import(record, decisions, user=self.user)

        existing.refresh_from_db()
        self.assertEqual(existing.chinese_translation, "新释义")
        self.assertFalse(existing.is_active)
        self.assertEqual(GlossaryEntry.objects.get(glossary=self.glossary, english_key="cable").chinese_translation, "线缆")
        self.assertFalse(GlossaryEntry.objects.filter(glossary=self.glossary, english_key="switch").exists())
        self.assertEqual(confirmed.result_summary["created"], 1)
        self.assertEqual(confirmed.result_summary["overwritten"], 1)
        self.assertEqual(confirmed.result_summary["skipped"], 1)
        self.assertEqual(len(confirmed.result_summary["rows"]), 3)

    def test_conflicting_duplicate_requires_selection_and_rolls_back(self):
        record = create_glossary_import(
            self.glossary,
            workbook_upload([[1, "Cable", "", "电缆"], [2, "cable", "", "线缆"]]),
            user=self.user,
        )

        with self.assertRaisesMessage(ValidationError, "尚未选择保留行"):
            confirm_glossary_import(record, {}, user=self.user)

        self.assertFalse(GlossaryEntry.objects.filter(glossary=self.glossary).exists())
        record.refresh_from_db()
        self.assertEqual(record.status, GlossaryImport.Status.PREVIEW)

    def test_changed_glossary_marks_preview_stale(self):
        record = create_glossary_import(
            self.glossary,
            workbook_upload([[1, "Router", "", "路由器"]]),
            user=self.user,
        )
        ProfessionalGlossary.objects.filter(pk=self.glossary.pk).update(
            updated_at=timezone.now() + timedelta(seconds=1)
        )

        with self.assertRaisesMessage(ValidationError, "预览后已发生变化"):
            confirm_glossary_import(record, {}, user=self.user)

        record.refresh_from_db()
        self.assertEqual(record.status, GlossaryImport.Status.STALE)
