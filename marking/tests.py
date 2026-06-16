import json
import shutil
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from assessments.models import Assessment, AssessmentModule
from competition_standards.models import CompetitionType, Project, StandardModule, TrainingCycle
from competitions.models import (
    Competition,
    CompetitionModule,
    CompetitionModuleStandardModuleMap,
    CompetitionPerson,
    CompetitionProject,
    CompetitionResult,
    Competitor,
    Member,
    MemberScope,
)
from core.constants import GROUP_COACH
from skilltrees.models import SkillNode, SkillTree

from .forms import MarkingAspectSkillNodeMapForm, MarkingSchemeImportForm, TARGET_ASSESSMENT_MODULE
from .models import (
    JudgementOption,
    MarkingAspect,
    MarkingAspectSkillNodeMap,
    MarkingParticipant,
    MarkingResult,
    MarkingSchemeImport,
)
from .parser import WorkbookParseError, parse_marking_workbook
from .services import create_scheme_from_upload, get_assessment_marking_score_map, import_result_package


User = get_user_model()
SAMPLE_WORKBOOK = Path(settings.BASE_DIR) / "media" / "CMP_marking_scheme_example.xlsx"
TEST_PRIVATE_MEDIA_ROOT = Path(settings.BASE_DIR) / ".marking-test-media"


def workbook_bytes(mutator=None):
    workbook = load_workbook(SAMPLE_WORKBOOK)
    if mutator is not None:
        mutator(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def uploaded_workbook(name="scheme.xlsx", mutator=None):
    content = SAMPLE_WORKBOOK.read_bytes() if mutator is None else workbook_bytes(mutator)
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@override_settings(PRIVATE_MEDIA_ROOT=TEST_PRIVATE_MEDIA_ROOT)
class MarkingWorkbookImportTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_PRIVATE_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        self.user = User.objects.create_superuser(
            username="marking-admin",
            password="testpass123",
            email="marking-admin@example.com",
        )
        coach_group = Group.objects.create(name=GROUP_COACH)
        self.user.groups.add(coach_group)
        self.participant = User.objects.create_user(
            username="participant-a",
            password="testpass123",
            first_name="学员甲",
        )
        self.competition_type = CompetitionType.objects.create(code="WSC-MARK", name="评分测试赛事")
        self.project = Project.objects.create(
            competition_type=self.competition_type,
            code="ITNSA-MARK",
            name="评分测试赛项",
        )
        self.module_set = self.project.get_or_create_default_standard_module_set()
        self.standard_module = StandardModule.objects.create(
            project=self.project,
            module_set=self.module_set,
            code="A",
            name="Linux environments",
        )
        self.training_cycle = TrainingCycle.objects.create(
            code="TC-MARK",
            name="评分测试周期",
            project=self.project,
            module_set=self.module_set,
            start_date=date(2026, 1, 1),
        )
        self.assessment = Assessment.objects.create(
            name="评分测试考核",
            training_cycle=self.training_cycle,
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 2),
        )
        self.assessment.participants.add(self.participant)
        self.assessment_module = AssessmentModule.objects.create(
            assessment=self.assessment,
            module=self.standard_module,
            responsible_coach=self.user,
            max_score=Decimal("3.60"),
        )

    def test_sample_workbook_parses_successfully(self):
        parsed = parse_marking_workbook(open(SAMPLE_WORKBOOK, "rb"))

        self.assertEqual(parsed.module_code, "A")
        self.assertEqual(parsed.module_name, "Linux environments")
        self.assertEqual(parsed.total_mark, Decimal("3.60"))
        self.assertEqual(len(parsed.subcriteria), 3)
        self.assertEqual(len(parsed.aspects), 10)
        judgement_aspect = next(aspect for aspect in parsed.aspects if aspect.aspect_type == "J")
        self.assertEqual(judgement_aspect.max_mark, Decimal("1.40"))
        self.assertEqual([option.score_value for option in judgement_aspect.judgement_options], [
            Decimal("0.00"),
            Decimal("1.00"),
            Decimal("2.00"),
            Decimal("3.00"),
        ])

    def test_missing_header_is_rejected(self):
        def mutate(workbook):
            workbook["Marking Scheme Import"].cell(row=6, column=11).value = "Max Points"

        with self.assertRaises(WorkbookParseError) as context:
            parse_marking_workbook(BytesIO(workbook_bytes(mutate)))

        self.assertIn("缺少必需表头", str(context.exception))
        self.assertIn("max mark", str(context.exception))

    def test_non_continuous_subcriterion_is_rejected(self):
        def mutate(workbook):
            workbook["Marking Scheme Import"].cell(row=11, column=1).value = "A3"

        with self.assertRaises(WorkbookParseError) as context:
            parse_marking_workbook(BytesIO(workbook_bytes(mutate)))

        self.assertIn("子评分项编号应为 A2", str(context.exception))

    def test_invalid_judgement_option_extra_column_is_rejected(self):
        def mutate(workbook):
            workbook["Marking Scheme Import"].cell(row=14, column=8).value = "不应填写"

        with self.assertRaises(WorkbookParseError) as context:
            parse_marking_workbook(BytesIO(workbook_bytes(mutate)))

        self.assertIn("J 分档行只能填写", str(context.exception))

    def test_valid_upload_creates_normalized_scheme_without_raw_rows(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook(),
            target=self.assessment_module,
            user=self.user,
        )

        self.assertEqual(scheme.module_code, "A")
        self.assertEqual(scheme.total_mark, Decimal("3.60"))
        self.assertEqual(scheme.subcriteria.count(), 3)
        self.assertEqual(scheme.aspects.count(), 10)
        self.assertEqual(JudgementOption.objects.filter(aspect__scheme=scheme).count(), 4)
        self.assertEqual(scheme.source_import.parse_summary["aspects_count"], 10)
        self.assertNotIn("raw_rows", scheme.source_import.parse_summary)

    def test_source_workbook_download_uses_authenticated_view(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("download-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("marking:scheme_source_download", args=[scheme.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response.headers["Content-Disposition"])

    def test_assessment_scheme_detail_is_filtered_by_target_access(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("access-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        outsider = User.objects.create_user(
            username="outsider",
            password="testpass123",
        )

        self.client.force_login(self.participant)
        participant_response = self.client.get(reverse("marking:scheme_detail", args=[scheme.pk]))
        self.assertEqual(participant_response.status_code, 200)

        self.client.force_login(outsider)
        outsider_response = self.client.get(reverse("marking:scheme_detail", args=[scheme.pk]))
        self.assertEqual(outsider_response.status_code, 404)

    def test_invalid_target_module_does_not_save_import_file(self):
        module_b = StandardModule.objects.create(
            project=self.project,
            module_set=self.module_set,
            code="B",
            name="Module B",
        )
        assessment_module_b = AssessmentModule.objects.create(
            assessment=self.assessment,
            module=module_b,
            max_score=Decimal("3.60"),
        )
        form = MarkingSchemeImportForm(
            data={
                "target_type": TARGET_ASSESSMENT_MODULE,
                "assessment_module": str(assessment_module_b.pk),
            },
            files={"file": uploaded_workbook()},
            user=self.user,
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(MarkingSchemeImport.objects.count(), 0)

    def test_json_result_package_imports_point_results_and_assessment_total(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("results-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        first_aspect = scheme.aspects.order_by("sort_order").first()
        payload = {
            "participants": [
                {
                    "username": self.participant.username,
                    "results": [
                        {
                            "aspect_code": first_aspect.code,
                            "score": "0.20",
                            "evidence": "CMP evidence",
                        }
                    ],
                },
                {
                    "external_identifier": "CMP-001",
                    "display_name": "外部选手",
                    "results": [
                        {
                            "aspect_code": first_aspect.code,
                            "score": "0.10",
                        }
                    ],
                },
            ]
        }

        result_import = import_result_package(
            scheme=scheme,
            uploaded_file=SimpleUploadedFile(
                "results.json",
                json.dumps(payload).encode("utf-8"),
                content_type="application/json",
            ),
            user=self.user,
        )

        self.assertEqual(result_import.summary["participant_count"], 2)
        self.assertEqual(MarkingResult.objects.count(), 2)
        self.assertTrue(MarkingParticipant.objects.filter(external_identifier="CMP-001").exists())
        score_map = get_assessment_marking_score_map(self.assessment)
        self.assertEqual(score_map[(self.participant.pk, self.assessment_module.pk)], Decimal("0.20"))

    def test_json_result_package_can_update_competition_official_result(self):
        competition = Competition.objects.create(
            competition_type=self.competition_type,
            code="WSC-MARK-2026",
            name="2026 评分测试赛",
            start_date=date(2026, 3, 1),
        )
        competition_project = CompetitionProject.objects.create(
            competition=competition,
            project=self.project,
        )
        competition_module = CompetitionModule.objects.create(
            competition_project=competition_project,
            code="A",
            name="Linux environments",
        )
        CompetitionModuleStandardModuleMap.objects.create(
            competition_module=competition_module,
            module=self.standard_module,
            is_primary=True,
        )
        member = Member.objects.create(
            name="中国",
            code="CN-MARK",
            level=MemberScope.INTERNATIONAL,
        )
        person = CompetitionPerson.objects.create(
            user=self.participant,
            name="选手甲",
            organization="测试单位",
        )
        competitor = Competitor.objects.create(
            person=person,
            competition_project=competition_project,
            member=member,
        )
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("competition-source.xlsx"),
            target=competition_module,
            user=self.user,
        )
        first_aspect = scheme.aspects.order_by("sort_order").first()
        payload = {
            "participants": [
                {
                    "competitor_id": competitor.pk,
                    "results": [{"aspect_code": first_aspect.code, "score": "0.20"}],
                    "official_result": {
                        "score_100": "91.50",
                        "score_700": "720.00",
                        "rank": 1,
                        "medal": "gold",
                    },
                }
            ]
        }

        import_result_package(
            scheme=scheme,
            uploaded_file=SimpleUploadedFile(
                "competition-results.json",
                json.dumps(payload).encode("utf-8"),
                content_type="application/json",
            ),
            user=self.user,
        )

        result = CompetitionResult.objects.get(competitor=competitor)
        self.assertEqual(result.score_100, Decimal("91.50"))
        self.assertEqual(result.score_700, Decimal("720.00"))
        self.assertEqual(result.rank, 1)
        self.assertEqual(result.medal, CompetitionResult.Medal.GOLD)

    def test_aspect_can_map_to_multiple_skill_nodes_with_one_primary(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("mapping-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        tree = SkillTree.objects.create(
            module=self.standard_module,
            name="Linux 技能树",
            version="v1",
            is_current=True,
        )
        node_a = SkillNode.objects.create(tree=tree, code="LIN-1", name="主机名配置")
        node_b = SkillNode.objects.create(tree=tree, code="LIN-2", name="网络配置")
        aspect = scheme.aspects.order_by("sort_order").first()

        MarkingAspectSkillNodeMap.objects.create(
            aspect=aspect,
            skill_node=node_a,
            is_primary=True,
            weight=Decimal("0.70"),
        )
        MarkingAspectSkillNodeMap.objects.create(
            aspect=aspect,
            skill_node=node_b,
            weight=Decimal("0.30"),
        )

        self.assertEqual(aspect.skill_node_mappings.count(), 2)
        self.assertEqual(aspect.skill_node_mappings.filter(is_primary=True).count(), 1)

    def test_aspect_rejects_non_skill_node_mapping(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("non-skill-mapping-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        tree = SkillTree.objects.create(
            module=self.standard_module,
            name="Linux 技能树",
            version="v1",
            is_current=True,
        )
        category = SkillNode.objects.create(
            tree=tree,
            code="LIN",
            name="Linux 基础",
            node_type=SkillNode.NodeType.CATEGORY,
        )
        aspect = scheme.aspects.order_by("sort_order").first()

        with self.assertRaises(ValidationError) as context:
            MarkingAspectSkillNodeMap.objects.create(aspect=aspect, skill_node=category)

        self.assertIn("评分点只能归类到技能点类型的节点", str(context.exception))

    def test_aspect_mapping_form_lists_only_active_skill_nodes(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("mapping-form-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        tree = SkillTree.objects.create(
            module=self.standard_module,
            name="Linux 技能树",
            version="v1",
            is_current=True,
        )
        active_skill = SkillNode.objects.create(tree=tree, code="LIN-1", name="主机名配置")
        inactive_skill = SkillNode.objects.create(
            tree=tree,
            code="LIN-2",
            name="停用技能点",
            is_active=False,
        )
        category = SkillNode.objects.create(
            tree=tree,
            code="LIN",
            name="Linux 基础",
            node_type=SkillNode.NodeType.CATEGORY,
        )
        task = SkillNode.objects.create(
            tree=tree,
            code="LIN-TASK",
            name="训练任务",
            node_type=SkillNode.NodeType.TASK,
        )
        aspect = scheme.aspects.order_by("sort_order").first()

        form = MarkingAspectSkillNodeMapForm(aspect=aspect)
        node_ids = set(form.fields["skill_node"].queryset.values_list("pk", flat=True))

        self.assertIn(active_skill.pk, node_ids)
        self.assertNotIn(inactive_skill.pk, node_ids)
        self.assertNotIn(category.pk, node_ids)
        self.assertNotIn(task.pk, node_ids)
        self.assertEqual(form.fields["skill_node"].label, "技能点")

    def test_aspect_rejects_skill_node_from_other_standard_module(self):
        scheme = create_scheme_from_upload(
            uploaded_file=uploaded_workbook("wrong-module-source.xlsx"),
            target=self.assessment_module,
            user=self.user,
        )
        module_b = StandardModule.objects.create(
            project=self.project,
            module_set=self.module_set,
            code="B",
            name="Module B",
        )
        other_tree = SkillTree.objects.create(module=module_b, name="B 树", version="v1")
        other_node = SkillNode.objects.create(tree=other_tree, code="B-1", name="其他节点")
        aspect = scheme.aspects.order_by("sort_order").first()

        with self.assertRaises(ValidationError):
            MarkingAspectSkillNodeMap.objects.create(aspect=aspect, skill_node=other_node)
