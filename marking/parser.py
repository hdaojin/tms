from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PARSER_VERSION = "cmp-single-module-v1"
MARKING_SHEET_NAME = "Marking Scheme Import"

EXPECTED_HEADERS = {
    "subcriterion_id": "sub criterion id",
    "subcriterion_name": "sub criterion name or description",
    "day_of_marking": "day of marking",
    "aspect_type": "aspect type m = meas j = judg",
    "aspect_description": "aspect - description",
    "judgement_score": "judg score",
    "extra_description": "extra aspect description (meas or judg) or judgement score description (judg only)",
    "requirement": "requirement (measurement only)",
    "wsos_section": "wsos section",
    "calculation_row": "calculation row (export only)",
    "max_mark": "max mark",
}

MODULE_HEADERS = {
    "module_code": "id",
    "module_name": "name",
    "module_mark": "mark",
}


class WorkbookParseError(ValueError):
    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))


@dataclass
class ParsedJudgementOption:
    score_value: Decimal
    description: str
    source_row_number: int
    sort_order: int


@dataclass
class ParsedAspect:
    code: str
    subcriterion_code: str
    aspect_type: str
    description: str
    command: str
    requirement: str
    wsos_section: str
    calculation_row: str
    max_mark: Decimal
    source_row_number: int
    sort_order: int
    judgement_options: list[ParsedJudgementOption] = field(default_factory=list)


@dataclass
class ParsedSubCriterion:
    code: str
    name: str
    day_of_marking: str
    source_row_number: int
    sort_order: int


@dataclass
class ParsedWorkbook:
    title: str
    module_code: str
    module_name: str
    module_mark: Decimal
    total_mark: Decimal
    subcriteria: list[ParsedSubCriterion]
    aspects: list[ParsedAspect]
    sheet_name: str = MARKING_SHEET_NAME
    parser_version: str = PARSER_VERSION

    def summary(self):
        return {
            "parser_version": self.parser_version,
            "sheet_name": self.sheet_name,
            "title": self.title,
            "module_code": self.module_code,
            "module_name": self.module_name,
            "module_mark": str(self.module_mark),
            "total_mark": str(self.total_mark),
            "subcriteria_count": len(self.subcriteria),
            "aspects_count": len(self.aspects),
            "measurement_count": sum(1 for aspect in self.aspects if aspect.aspect_type == "M"),
            "judgement_count": sum(1 for aspect in self.aspects if aspect.aspect_type == "J"),
        }


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_header(value: Any) -> str:
    return normalize_text(value).lower()


def is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(normalize_text(value) == "" for value in row)


def to_decimal(value: Any, *, row_number: int, field_label: str) -> Decimal:
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if value is None or normalize_text(value) == "":
        raise WorkbookParseError([f"第 {row_number} 行：{field_label} 不能为空。"])
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        raise WorkbookParseError([f"第 {row_number} 行：{field_label} 必须是可计算的数值。"])


def calculate_file_sha256(file_obj) -> str:
    digest = hashlib.sha256()
    position = None
    if hasattr(file_obj, "tell") and hasattr(file_obj, "seek"):
        position = file_obj.tell()
        file_obj.seek(0)
    for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
        digest.update(chunk)
    if position is not None:
        file_obj.seek(position)
    return digest.hexdigest()


def calculate_path_sha256(path: str | Path) -> str:
    with open(path, "rb") as file:
        return calculate_file_sha256(file)


def _build_mapping(row: tuple[Any, ...], expected: dict[str, str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [normalize_header(value) for value in row]
    for field, header in expected.items():
        if header in normalized:
            mapping[field] = normalized.index(header)
    return mapping


def _find_header_row(values: list[tuple[Any, ...]], expected: dict[str, str]) -> tuple[int, dict[str, int]]:
    best_mapping: dict[str, int] = {}
    for index, row in enumerate(values, start=1):
        mapping = _build_mapping(row, expected)
        if set(mapping) == set(expected):
            return index, mapping
        if len(mapping) > len(best_mapping):
            best_mapping = mapping
    if best_mapping:
        missing = "、".join(header for field, header in expected.items() if field not in best_mapping)
        raise WorkbookParseError([f"缺少必需表头或表头拼写不符：{missing}。"])
    required = "、".join(expected.values())
    raise WorkbookParseError([f"未找到完整表头行，必须包含：{required}。"])


def _value(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    index = mapping[field]
    if index >= len(row):
        return None
    return row[index]


def _text(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> str:
    return normalize_text(_value(row, mapping, field))


def _unexpected_fields(row: tuple[Any, ...], mapping: dict[str, int], allowed: set[str]) -> list[str]:
    unexpected = []
    for field, index in mapping.items():
        if field in allowed:
            continue
        if index < len(row) and normalize_text(row[index]):
            unexpected.append(EXPECTED_HEADERS[field])
    return unexpected


def _is_total_row(row: tuple[Any, ...], mapping: dict[str, int]) -> bool:
    return _text(row, mapping, "calculation_row").lower() == "total marks:"


def _is_subcriterion_row(row: tuple[Any, ...], mapping: dict[str, int]) -> bool:
    return bool(_text(row, mapping, "subcriterion_id"))


def _is_aspect_row(row: tuple[Any, ...], mapping: dict[str, int]) -> bool:
    return _text(row, mapping, "aspect_type").upper() in {"M", "J"}


def _parse_module_header(values: list[tuple[Any, ...]]) -> tuple[str, str, str, Decimal]:
    header_row_number, mapping = _find_header_row(values, MODULE_HEADERS)
    if header_row_number >= len(values):
        raise WorkbookParseError([f"第 {header_row_number} 行：模块表头后缺少模块数据行。"])
    data_row_number = header_row_number + 1
    data_row = values[data_row_number - 1]
    module_code = normalize_text(_value(data_row, mapping, "module_code"))
    module_name = normalize_text(_value(data_row, mapping, "module_name"))
    module_mark = to_decimal(_value(data_row, mapping, "module_mark"), row_number=data_row_number, field_label="Mark")
    errors = []
    if not module_code:
        errors.append(f"第 {data_row_number} 行：ID 不能为空。")
    if not module_name:
        errors.append(f"第 {data_row_number} 行：Name 不能为空。")
    if errors:
        raise WorkbookParseError(errors)
    title = normalize_text(values[0][0]) if values and values[0] else ""
    return title, module_code, module_name, module_mark


def parse_marking_workbook(file_or_path) -> ParsedWorkbook:
    try:
        if hasattr(file_or_path, "seek"):
            file_or_path.seek(0)
        workbook = load_workbook(file_or_path, data_only=True, read_only=True)
    except Exception as exc:
        raise WorkbookParseError([f"无法读取 xlsx 文件：{exc}"])

    if MARKING_SHEET_NAME not in workbook.sheetnames:
        raise WorkbookParseError([f"未找到工作表：{MARKING_SHEET_NAME}。"])

    worksheet = workbook[MARKING_SHEET_NAME]
    values = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    if not values:
        raise WorkbookParseError(["评分表为空。"])

    title, module_code, module_name, module_mark = _parse_module_header(values)
    header_row_number, mapping = _find_header_row(values, EXPECTED_HEADERS)

    errors: list[str] = []
    subcriteria: list[ParsedSubCriterion] = []
    aspects: list[ParsedAspect] = []
    current_subcriterion: ParsedSubCriterion | None = None
    expected_subcriterion_number = 1
    aspect_order = 0
    aspect_count_by_subcriterion: dict[str, int] = {}
    total_mark: Decimal | None = None
    row_index = header_row_number + 1

    while row_index <= len(values):
        row_number = row_index
        row = values[row_index - 1]
        if is_blank_row(row):
            row_index += 1
            continue
        if _is_total_row(row, mapping):
            try:
                total_mark = to_decimal(_value(row, mapping, "max_mark"), row_number=row_number, field_label="Total Marks")
            except WorkbookParseError as exc:
                errors.extend(exc.errors)
            break
        if _is_subcriterion_row(row, mapping):
            unexpected = _unexpected_fields(row, mapping, {"subcriterion_id", "subcriterion_name", "day_of_marking"})
            if unexpected:
                errors.append(f"第 {row_number} 行：子评分项标题行只能填写前三个语义字段，不能填写：{'、'.join(unexpected)}。")
            code = _text(row, mapping, "subcriterion_id")
            name = _text(row, mapping, "subcriterion_name")
            day = _text(row, mapping, "day_of_marking")
            match = re.fullmatch(rf"{re.escape(module_code)}(\d+)", code)
            if not match:
                errors.append(f"第 {row_number} 行：子评分项编号必须形如 {module_code}1、{module_code}2。")
            else:
                actual_number = int(match.group(1))
                if actual_number != expected_subcriterion_number:
                    errors.append(
                        f"第 {row_number} 行：子评分项编号应为 {module_code}{expected_subcriterion_number}，实际为 {code}。"
                    )
                expected_subcriterion_number += 1
            if not name:
                errors.append(f"第 {row_number} 行：Sub Criterion Name or Description 不能为空。")
            if not day:
                errors.append(f"第 {row_number} 行：Day of Marking 不能为空。")
            current_subcriterion = ParsedSubCriterion(
                code=code,
                name=name,
                day_of_marking=day,
                source_row_number=row_number,
                sort_order=len(subcriteria),
            )
            subcriteria.append(current_subcriterion)
            row_index += 1
            continue

        if _is_aspect_row(row, mapping):
            if current_subcriterion is None:
                errors.append(f"第 {row_number} 行：评分点必须位于某个子评分项下。")
                row_index += 1
                continue
            aspect_type = _text(row, mapping, "aspect_type").upper()
            allowed = {
                "aspect_type",
                "aspect_description",
                "extra_description",
                "requirement",
                "wsos_section",
                "calculation_row",
                "max_mark",
            }
            unexpected = _unexpected_fields(row, mapping, allowed)
            if unexpected:
                errors.append(f"第 {row_number} 行：评分点行不能填写：{'、'.join(unexpected)}。")
            description = _text(row, mapping, "aspect_description")
            command = _text(row, mapping, "extra_description")
            requirement = _text(row, mapping, "requirement")
            if not description:
                errors.append(f"第 {row_number} 行：Aspect - Description 不能为空。")
            if not command:
                errors.append(f"第 {row_number} 行：命令或操作说明不能为空。")
            if aspect_type == "M" and not requirement:
                errors.append(f"第 {row_number} 行：M 类型评分点的 Requirement 不能为空。")
            try:
                max_mark = to_decimal(_value(row, mapping, "max_mark"), row_number=row_number, field_label="Max Mark")
            except WorkbookParseError as exc:
                errors.extend(exc.errors)
                max_mark = Decimal("0.00")

            aspect_number = aspect_count_by_subcriterion.get(current_subcriterion.code, 0) + 1
            aspect_count_by_subcriterion[current_subcriterion.code] = aspect_number
            aspect = ParsedAspect(
                code=f"{current_subcriterion.code}.{aspect_number}",
                subcriterion_code=current_subcriterion.code,
                aspect_type=aspect_type,
                description=description,
                command=command,
                requirement=requirement,
                wsos_section=_text(row, mapping, "wsos_section"),
                calculation_row=_text(row, mapping, "calculation_row"),
                max_mark=max_mark,
                source_row_number=row_number,
                sort_order=aspect_order,
            )
            aspect_order += 1

            if aspect_type == "J":
                option_scores = []
                option_extra_mark_total = Decimal("0.00")
                for option_order, expected_score in enumerate((0, 1, 2, 3), start=1):
                    option_row_index = row_index + option_order
                    if option_row_index > len(values):
                        errors.append(f"第 {row_number} 行：J 类型评分点缺少 {expected_score} 分分档行。")
                        continue
                    option_row_number = option_row_index
                    option_row = values[option_row_number - 1]
                    allowed_option_fields = {"judgement_score", "extra_description", "calculation_row", "max_mark"}
                    unexpected = _unexpected_fields(option_row, mapping, allowed_option_fields)
                    if unexpected:
                        errors.append(
                            f"第 {option_row_number} 行：J 分档行只能填写 Judg Score、分档描述和导出分值列，"
                            f"不能填写：{'、'.join(unexpected)}。"
                        )
                    try:
                        score_value = to_decimal(
                            _value(option_row, mapping, "judgement_score"),
                            row_number=option_row_number,
                            field_label="Judg Score",
                        )
                    except WorkbookParseError as exc:
                        errors.extend(exc.errors)
                        score_value = Decimal("-1.00")
                    option_mark_value = _value(option_row, mapping, "max_mark")
                    if normalize_text(option_mark_value) != "":
                        try:
                            option_extra_mark_total += to_decimal(
                                option_mark_value,
                                row_number=option_row_number,
                                field_label="Max Mark",
                            )
                        except WorkbookParseError as exc:
                            errors.extend(exc.errors)
                    option_scores.append(score_value)
                    description_text = _text(option_row, mapping, "extra_description")
                    if not description_text:
                        errors.append(f"第 {option_row_number} 行：J 分档说明不能为空。")
                    aspect.judgement_options.append(
                        ParsedJudgementOption(
                            score_value=score_value,
                            description=description_text,
                            source_row_number=option_row_number,
                            sort_order=option_order - 1,
                        )
                    )
                if option_scores != [Decimal("0.00"), Decimal("1.00"), Decimal("2.00"), Decimal("3.00")]:
                    errors.append(f"第 {row_number} 行：J 类型评分点必须按 0、1、2、3 四档连续填写。")
                aspect.max_mark += option_extra_mark_total
                row_index += 5
            else:
                row_index += 1
            aspects.append(aspect)
            continue

        errors.append(f"第 {row_number} 行：无法识别为子评分项、评分点或总分行。")
        row_index += 1

    if not subcriteria:
        errors.append("评分表至少需要一个子评分项。")
    if not aspects:
        errors.append("评分表至少需要一个评分点。")
    if total_mark is None:
        errors.append("未找到 Total Marks: 总分行。")
        total_mark = Decimal("0.00")

    aspect_total = sum((aspect.max_mark for aspect in aspects), Decimal("0.00")).quantize(Decimal("0.01"))
    if aspect_total != total_mark:
        errors.append(f"评分点分值总和为 {aspect_total}，与 Total Marks {total_mark} 不一致。")
    if total_mark != module_mark:
        errors.append(f"Total Marks {total_mark} 与模块 Mark {module_mark} 不一致。")

    if errors:
        raise WorkbookParseError(errors)

    return ParsedWorkbook(
        title=title,
        module_code=module_code,
        module_name=module_name,
        module_mark=module_mark,
        total_mark=total_mark,
        subcriteria=subcriteria,
        aspects=aspects,
    )
