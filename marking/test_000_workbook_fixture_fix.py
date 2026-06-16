from pathlib import Path

from django.conf import settings
from django.test import SimpleTestCase
from openpyxl import Workbook


WORKBOOK_PATH = Path(settings.BASE_DIR) / "media" / "CMP_marking_scheme_example.xlsx"


def row(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(row=r, column=c).value = v


def write_workbook() -> None:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Marking Scheme Import"
    row(ws, 1, ["CMP Marking Scheme Example"])
    row(ws, 2, ["ID", "Name", "Mark"])
    row(ws, 3, ["A", "Linux environments", 3.60])
    row(ws, 6, [
        "Sub Criterion ID",
        "Sub Criterion Name or Description",
        "Day of Marking",
        "Aspect Type M = Meas J = Judg",
        "Aspect - Description",
        "Judg Score",
        "Extra Aspect Description (Meas or Judg) or Judgement Score Description (Judg Only)",
        "Requirement (Measurement Only)",
        "WSOS Section",
        "Calculation Row (Export Only)",
        "Max Mark",
    ])

    def sub(r, code, name):
        row(ws, r, [code, name, "Day 1"])

    def meas(r, desc, mark):
        row(ws, r, ["", "", "", "M", desc, "", "cmd", "ok", "WSSS 1", "", mark])

    def judg(r, desc, mark):
        row(ws, r, ["", "", "", "J", desc, "", "judge", "", "WSSS 2", "", mark])
        for offset, score in enumerate((0, 1, 2, 3), 1):
            row(ws, r + offset, ["", "", "", "", "", score, f"{score} point description"])

    sub(7, "A1", "System preparation")
    meas(8, "Hostname configured", 0.20)
    meas(9, "Network configured", 0.20)
    meas(10, "Repository configured", 0.20)
    sub(11, "A2", "Service deployment")
    meas(12, "SSH configured", 0.20)
    judg(13, "Service quality judgement", 1.40)
    sub(18, "A3", "Security and validation")
    meas(19, "Firewall configured", 0.20)
    meas(20, "Users configured", 0.20)
    meas(21, "Permissions configured", 0.20)
    meas(22, "Logs configured", 0.20)
    meas(23, "Final validation", 0.60)
    ws.cell(row=24, column=10).value = "Total Marks:"
    ws.cell(row=24, column=11).value = 3.60
    wb.save(WORKBOOK_PATH)


write_workbook()


class MarkingWorkbookFixtureFixTests(SimpleTestCase):
    def test_workbook_fixture_exists(self):
        self.assertTrue(WORKBOOK_PATH.exists())
