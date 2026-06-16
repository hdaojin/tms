from pathlib import Path

from django.conf import settings
from django.test import SimpleTestCase
from openpyxl import Workbook


WORKBOOK_PATH = Path(settings.BASE_DIR) / "media" / "CMP_marking_scheme_example.xlsx"


def make_workbook() -> None:
    if WORKBOOK_PATH.exists():
        return

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Marking Scheme Import"
    ws.cell(row=1, column=1).value = "CMP Marking Scheme Example"
    ws.cell(row=2, column=1).value = "ID"
    ws.cell(row=2, column=2).value = "Name"
    ws.cell(row=2, column=3).value = "Mark"
    ws.cell(row=3, column=1).value = "A"
    ws.cell(row=3, column=2).value = "Linux environments"
    ws.cell(row=3, column=3).value = 3.60

    headers = [
        "Sub Criterion ID",
        "Sub Criterion Name or Description",
        "Day of Marking",
        "Aspect Type M = Meas J = Judg",
        "Aspect - Description",
        "Judg Score",
        "Extra Aspect Description (Meas or Judg) or Judgement Score Description (Judg Only)",
        "Requirement (Measurement Only)",
        "WSOS Section",
        "Calculation Row (Export Only)",
        "Max Mark",
    ]
    for column, header in enumerate(headers, start=1):
        ws.cell(row=6, column=column).value = header

    rows = [
        (7, "A1", "System preparation", "", "", "", "", "", "", "", ""),
        (8, "", "", "M", "Hostname configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (9, "", "", "M", "Network configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (10, "", "", "M", "Repository configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (11, "A2", "Service deployment", "", "", "", "", "", "", "", ""),
        (12, "", "", "M", "SSH configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (13, "", "", "J", "Service quality judgement", "", "Evaluate service", "", "WSSS 2", "", 1.40),
        (14, "", "", "", "", 0, "0 point description", "", "", "", ""),
        (15, "", "", "", "", 1, "1 point description", "", "", "", ""),
        (16, "", "", "", "", 2, "2 point description", "", "", "", ""),
        (17, "", "", "", "", 3, "3 point description", "", "", "", ""),
        (18, "A3", "Security and validation", "", "", "", "", "", "", "", ""),
        (19, "", "", "M", "Firewall configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (20, "", "", "M", "Users configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (21, "", "", "M", "Permissions configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (22, "", "", "M", "Logs configured", "", "Run command", "Expected result", "WSSS 1", "", 0.20),
        (23, "", "", "M", "Final validation", "", "Run command", "Expected result", "WSSS 1", "", 0.60),
    ]
    for row in rows:
        row_number = row[0]
        for column, value in enumerate(row[1:], start=1):
            ws.cell(row=row_number, column=column).value = value

    ws.cell(row=24, column=10).value = "Total Marks:"
    ws.cell(row=24, column=11).value = 3.60
    wb.save(WORKBOOK_PATH)


make_workbook()


class MarkingWorkbookFixtureTests(SimpleTestCase):
    def test_workbook_fixture_exists(self):
        self.assertTrue(WORKBOOK_PATH.exists())
