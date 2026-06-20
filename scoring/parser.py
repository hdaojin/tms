from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .registry import CMP_SINGLE_MODULE_V1


PARSER_VERSION = CMP_SINGLE_MODULE_V1
MARKING_SHEET_NAME = "Marking Scheme Import"
CENT = Decimal("0.01")
EXPECTED_DETAIL_HEADERS = {
    "subcriterion_code": "Sub Criterion ID",
    "subcriterion_name": "Sub Criterion Name or Description",
    "day_of_marking": "Day of Marking",
    "aspect_type": "Aspect Type M = Meas J = Judg",
    "description": "Aspect - Description",
    "judgement_score": "Judg Score",
    "extra_description": "Extra Aspect Description (Meas or Judg) OR Judgement Score Description (Judg only)",
    "requirement": "Requirement (Measurement Only)",
    "wsos_section": "WSOS Section",
    "calculation_row": "Calculation Row (Export only)",
    "max_mark": "Max Mark",
}
EXPECTED_MODULE_HEADERS = {
    "module_code": "ID",
    "module_name": "Name",
    "module_mark": "Mark",
}
DETAIL_HEADER_LOOKUP = {"": ""} | {
    re.sub(r"\s+", " ", value).strip().lower(): key for key, value in EXPECTED_DETAIL_HEADERS.items()
}
MODULE_HEADER_LOOKUP = {"": ""} | {
    re.sub(r"\s+", " ", value).strip().lower(): key for key, value in EXPECTED_MODULE_HEADERS.items()
}


class WorkbookParseError(ValueError):
    def __init__(self, message_or_errors: str | list[str]):
        self.errors = [message_or_errors] if isinstance(message_or_errors, str) else message_or_errors
        super().__init__("\n".join(self.errors))


@dataclass(frozen=True)
class ParsedJudgementOption:
    score_value: Decimal
    description: str
    source_row_number: int
    order: int = 0

    def as_dict(self):
        return {
            "score_value": str(self.score_value),
            "description": self.description,
            "source_row_number": self.source_row_number,
            "order": self.order,
        }


@dataclass(frozen=True)
class ParsedAspect:
    code: str
    aspect_type: str
    description: str
    command: str
    requirement: str
    calculation_row: str
    max_mark: Decimal
    source_row_number: int
    order: int = 0
    judgement_options: list[ParsedJudgementOption] = field(default_factory=list)

    def as_dict(self):
        return {
            "code": self.code,
            "aspect_type": self.aspect_type,
            "description": self.description,
            "command": self.command,
            "requirement": self.requirement,
            "calculation_row": self.calculation_row,
            "max_mark": str(self.max_mark),
            "source_row_number": self.source_row_number,
            "order": self.order,
            "judgement_options": [option.as_dict() for option in self.judgement_options],
        }


@dataclass(frozen=True)
class ParsedSubCriterion:
    code: str
    name: str
    day_of_marking: str
    order: int = 0
    aspects: list[ParsedAspect] = field(default_factory=list)

    def as_dict(self):
        return {
            "code": self.code,
            "name": self.name,
            "day_of_marking": self.day_of_marking,
            "order": self.order,
            "aspects": [aspect.as_dict() for aspect in self.aspects],
        }


@dataclass(frozen=True)
class ParsedWorkbook:
    module_code: str
    module_name: str
    module_mark: Decimal
    title: str
    total_mark: Decimal
    subcriteria: list[ParsedSubCriterion]
    raw_snapshot: dict[str, Any]
    field_mapping: dict[str, Any]
    validation_report: dict[str, Any]

    def as_payload(self):
        aspects = [aspect for subcriterion in self.subcriteria for aspect in subcriterion.aspects]
        return {
            "parser_version": PARSER_VERSION,
            "title": self.title,
            "module_code": self.module_code,
            "module_name": self.module_name,
            "module_mark": str(self.module_mark),
            "total_mark": str(self.total_mark),
            "subcriteria_count": len(self.subcriteria),
            "aspects_count": len(aspects),
            "measurement_count": sum(1 for aspect in aspects if aspect.aspect_type == "M"),
            "judgement_count": sum(1 for aspect in aspects if aspect.aspect_type == "J"),
            "subcriteria": [subcriterion.as_dict() for subcriterion in self.subcriteria],
        }


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip().lower()


def _normalize_content(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"[^\S\n]+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(lines).strip()


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


def _open_workbooks(file_or_path):
    try:
        if hasattr(file_or_path, "seek"):
            file_or_path.seek(0)
        formula_workbook = load_workbook(file_or_path, data_only=False)
        if hasattr(file_or_path, "seek"):
            file_or_path.seek(0)
        value_workbook = load_workbook(file_or_path, data_only=True)
    except Exception as exc:
        raise WorkbookParseError(f"无法读取 xlsx 文件：{exc}") from exc
    return formula_workbook, value_workbook


def _build_raw_snapshot(formula_sheet, value_sheet) -> dict[str, Any]:
    rows = []
    for formula_row, value_row in zip(
        formula_sheet.iter_rows(min_row=1, max_row=formula_sheet.max_row, max_col=formula_sheet.max_column),
        value_sheet.iter_rows(min_row=1, max_row=formula_sheet.max_row, max_col=formula_sheet.max_column),
        strict=True,
    ):
        cells = []
        for formula_cell, value_cell in zip(formula_row, value_row, strict=True):
            is_formula = isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
            value = value_cell.value if is_formula else formula_cell.value
            cells.append(
                {
                    "row": formula_cell.row,
                    "column": formula_cell.column,
                    "column_letter": get_column_letter(formula_cell.column),
                    "coordinate": formula_cell.coordinate,
                    "value": _json_value(value),
                    "formula": formula_cell.value if is_formula else "",
                    "normalized_text": _normalize_content(value),
                }
            )
        rows.append({"row_number": formula_row[0].row, "cells": cells})
    return {"sheet_name": formula_sheet.title, "max_row": formula_sheet.max_row, "max_column": formula_sheet.max_column, "rows": rows}


def _cell(row: dict[str, Any], column_index: int) -> dict[str, Any]:
    return row["cells"][column_index - 1]


def _cell_text(row: dict[str, Any], column_index: int) -> str:
    return _cell(row, column_index)["normalized_text"]


def _non_empty_cells(row: dict[str, Any]) -> list[dict[str, Any]]:
    return [cell for cell in row["cells"] if cell["normalized_text"]]


def _is_blank_row(row: dict[str, Any]) -> bool:
    return not _non_empty_cells(row)


def _header_mapping(row: dict[str, Any], lookup: dict[str, str]) -> tuple[dict[str, int], list[str], list[str]]:
    mapping = {}
    unexpected = []
    duplicates = []
    for cell in _non_empty_cells(row):
        key = lookup.get(_normalize_header(cell["value"]))
        if not key:
            unexpected.append(cell["coordinate"])
            continue
        if key in mapping:
            duplicates.append(EXPECTED_DETAIL_HEADERS.get(key, EXPECTED_MODULE_HEADERS.get(key, key)))
            continue
        mapping[key] = cell["column"]
    return mapping, unexpected, duplicates


def _find_header_row(
    rows: list[dict[str, Any]],
    lookup: dict[str, str],
    expected: dict[str, str],
    start_index: int = 0,
) -> tuple[int | None, dict[str, int], list[str]]:
    best_index = None
    best_mapping = {}
    errors = []
    for index, row in enumerate(rows[start_index:], start=start_index):
        mapping, unexpected, duplicates = _header_mapping(row, lookup)
        if len(mapping) > len(best_mapping):
            best_index = index
            best_mapping = mapping
        if set(mapping) == set(expected) and not unexpected and not duplicates:
            return index, mapping, []
    if best_index is None:
        errors.append("未找到表头行。")
    else:
        missing = [label for key, label in expected.items() if key not in best_mapping]
        if missing:
            errors.append(f"第 {rows[best_index]['row_number']} 行缺少表头：{'、'.join(missing)}。")
    return None, best_mapping, errors


def _field_text(row: dict[str, Any], mapping: dict[str, int], field: str) -> str:
    column = mapping.get(field)
    return _cell_text(row, column) if column else ""


def _field_cell(row: dict[str, Any], mapping: dict[str, int], field: str) -> dict[str, Any] | None:
    column = mapping.get(field)
    return _cell(row, column) if column else None


def _row_unexpected_errors(
    row: dict[str, Any],
    mapping: dict[str, int],
    allowed_fields: set[str],
    field_labels: dict[str, str],
    row_label: str,
) -> list[str]:
    allowed_columns = {mapping[field] for field in allowed_fields if field in mapping}
    mapped_columns = set(mapping.values())
    errors = []
    for cell in _non_empty_cells(row):
        if cell["column"] not in mapped_columns:
            errors.append(f"第 {row['row_number']} 行：{row_label}存在模板外内容（{cell['coordinate']}）。")
        elif cell["column"] not in allowed_columns:
            field = next(key for key, column in mapping.items() if column == cell["column"])
            errors.append(f"第 {row['row_number']} 行：{row_label}不能填写 {field_labels[field]}。")
    return errors


def _decimal_from_cell(cell: dict[str, Any] | None, *, row_number: int, field_label: str, required: bool = False):
    if cell is None or cell["normalized_text"] == "":
        if required:
            return None, f"第 {row_number} 行：{field_label} 不能为空。"
        return Decimal("0.00"), ""
    value = cell["value"]
    if cell.get("formula") and value in (None, ""):
        return None, f"第 {row_number} 行：{field_label} 公式没有可读取的缓存数值。"
    try:
        return Decimal(str(value)).quantize(CENT), ""
    except (InvalidOperation, ValueError):
        return None, f"第 {row_number} 行：{field_label} 必须是可计算的数值。"


def _is_total_row(row: dict[str, Any], mapping: dict[str, int]) -> bool:
    return _normalize_header(_field_text(row, mapping, "calculation_row")) == "total marks:"


def _is_subcriterion_row(row: dict[str, Any], mapping: dict[str, int]) -> bool:
    return bool(_field_text(row, mapping, "subcriterion_code")) and not _field_text(row, mapping, "aspect_type")


def _is_aspect_row(row: dict[str, Any], mapping: dict[str, int]) -> bool:
    return _field_text(row, mapping, "aspect_type").upper() in {"M", "J"}


def _is_judgement_option_row(row: dict[str, Any], mapping: dict[str, int]) -> bool:
    if _is_total_row(row, mapping) or _is_subcriterion_row(row, mapping) or _is_aspect_row(row, mapping):
        return False
    return bool(_field_text(row, mapping, "judgement_score") or _field_text(row, mapping, "extra_description"))


def _mapping_payload(mapping: dict[str, int], raw_snapshot: dict[str, Any], row_index: int, labels: dict[str, str]):
    header_row = raw_snapshot["rows"][row_index]
    return {
        "row_number": header_row["row_number"],
        "fields": {
            key: {
                "label": labels[key],
                "column": column,
                "column_letter": get_column_letter(column),
                "header_value": _cell(header_row, column)["value"],
            }
            for key, column in mapping.items()
        },
    }


def _find_title(rows: list[dict[str, Any]], errors: list[str]) -> str:
    if not rows or _is_blank_row(rows[0]):
        errors.append("第 1 行：标题不能为空。")
        return ""
    non_empty = _non_empty_cells(rows[0])
    if len(non_empty) > 1:
        errors.append("第 1 行：标题行只能填写一个标题单元格。")
    return non_empty[0]["normalized_text"]


def parse_marking_workbook(file_or_path, *, expected_module_code: str = "") -> ParsedWorkbook:
    formula_workbook, value_workbook = _open_workbooks(file_or_path)
    errors: list[str] = []
    warnings: list[str] = []
    if MARKING_SHEET_NAME not in formula_workbook.sheetnames:
        raise WorkbookParseError(f"未找到工作表：{MARKING_SHEET_NAME}。")

    formula_sheet = formula_workbook[MARKING_SHEET_NAME]
    value_sheet = value_workbook[MARKING_SHEET_NAME]
    raw_snapshot = _build_raw_snapshot(formula_sheet, value_sheet)
    rows = raw_snapshot["rows"]
    title = _find_title(rows, errors)

    module_header_index, module_mapping, module_errors = _find_header_row(
        rows,
        MODULE_HEADER_LOOKUP,
        EXPECTED_MODULE_HEADERS,
        start_index=1,
    )
    errors.extend(module_errors)
    module_code = ""
    module_name = ""
    module_mark = Decimal("0.00")
    module_data_index = None
    if module_header_index is not None:
        module_data_index = module_header_index + 1
        while module_data_index < len(rows) and _is_blank_row(rows[module_data_index]):
            module_data_index += 1
        if module_data_index >= len(rows):
            errors.append(f"第 {rows[module_header_index]['row_number']} 行：模块表头后缺少模块数据行。")
        else:
            module_row = rows[module_data_index]
            errors.extend(
                _row_unexpected_errors(
                    module_row,
                    module_mapping,
                    {"module_code", "module_name", "module_mark"},
                    EXPECTED_MODULE_HEADERS,
                    "模块定义行",
                )
            )
            module_code = _field_text(module_row, module_mapping, "module_code")
            module_name = _field_text(module_row, module_mapping, "module_name")
            module_mark, error = _decimal_from_cell(
                _field_cell(module_row, module_mapping, "module_mark"),
                row_number=module_row["row_number"],
                field_label="Mark",
                required=True,
            )
            if error:
                errors.append(error)
                module_mark = Decimal("0.00")
            if not module_code:
                errors.append(f"第 {module_row['row_number']} 行：ID 不能为空。")
            if not module_name:
                errors.append(f"第 {module_row['row_number']} 行：Name 不能为空。")
            if expected_module_code and module_code and module_code != expected_module_code:
                errors.append(f"第 {module_row['row_number']} 行：模块 ID 必须与事件模块一致：{expected_module_code}。")

    detail_header_index, detail_mapping, detail_errors = _find_header_row(
        rows,
        DETAIL_HEADER_LOOKUP,
        EXPECTED_DETAIL_HEADERS,
        start_index=(module_data_index + 1 if module_data_index is not None else 0),
    )
    errors.extend(detail_errors)

    if module_header_index is not None:
        for extra_row in rows[1:module_header_index]:
            if not _is_blank_row(extra_row):
                errors.append(f"第 {extra_row['row_number']} 行：标题与模块定义之间不能填写业务内容。")
    if module_data_index is not None and detail_header_index is not None:
        for extra_row in rows[module_data_index + 1 : detail_header_index]:
            if not _is_blank_row(extra_row):
                errors.append(f"第 {extra_row['row_number']} 行：模块定义与评分细则表头之间不能填写业务内容。")

    subcriteria: list[ParsedSubCriterion] = []
    current_subcriterion: ParsedSubCriterion | None = None
    aspect_count_by_subcriterion: dict[str, int] = {}
    aspect_order = 0
    total_mark: Decimal | None = None
    total_row_seen = False

    if detail_header_index is not None:
        row_index = detail_header_index + 1
        expected_subcriterion_number = 1
        while row_index < len(rows):
            row = rows[row_index]
            row_number = row["row_number"]
            if _is_blank_row(row):
                row_index += 1
                continue
            if _is_total_row(row, detail_mapping):
                total_row_seen = True
                errors.extend(
                    _row_unexpected_errors(
                        row,
                        detail_mapping,
                        {"calculation_row", "max_mark"},
                        EXPECTED_DETAIL_HEADERS,
                        "总分行",
                    )
                )
                total_mark, error = _decimal_from_cell(
                    _field_cell(row, detail_mapping, "max_mark"),
                    row_number=row_number,
                    field_label="Total Marks",
                    required=True,
                )
                if error:
                    errors.append(error)
                    total_mark = Decimal("0.00")
                for later_row in rows[row_index + 1 :]:
                    if not _is_blank_row(later_row):
                        errors.append(f"第 {later_row['row_number']} 行：Total Marks 后不能再填写内容。")
                break

            if _is_subcriterion_row(row, detail_mapping):
                errors.extend(
                    _row_unexpected_errors(
                        row,
                        detail_mapping,
                        {"subcriterion_code", "subcriterion_name", "day_of_marking"},
                        EXPECTED_DETAIL_HEADERS,
                        "子评分项标题行",
                    )
                )
                code = _field_text(row, detail_mapping, "subcriterion_code")
                name = _field_text(row, detail_mapping, "subcriterion_name")
                if not re.fullmatch(rf"{re.escape(module_code)}\d+", code or ""):
                    errors.append(f"第 {row_number} 行：子评分项编号必须形如 {module_code}1、{module_code}2。")
                else:
                    expected_code = f"{module_code}{expected_subcriterion_number}"
                    if code != expected_code:
                        errors.append(f"第 {row_number} 行：子评分项编号应为 {expected_code}，实际为 {code}。")
                    expected_subcriterion_number += 1
                if not name:
                    errors.append(f"第 {row_number} 行：Sub Criterion Name or Description 不能为空。")
                current_subcriterion = ParsedSubCriterion(
                    code=code,
                    name=name,
                    day_of_marking=_field_text(row, detail_mapping, "day_of_marking"),
                    order=len(subcriteria),
                    aspects=[],
                )
                subcriteria.append(current_subcriterion)
                row_index += 1
                continue

            if _is_aspect_row(row, detail_mapping):
                if current_subcriterion is None:
                    errors.append(f"第 {row_number} 行：评分点必须位于某个子评分项下。")
                    row_index += 1
                    continue
                aspect_type = _field_text(row, detail_mapping, "aspect_type").upper()
                allowed = {
                    "aspect_type",
                    "description",
                    "extra_description",
                    "requirement",
                    "wsos_section",
                    "calculation_row",
                    "max_mark",
                }
                errors.extend(
                    _row_unexpected_errors(row, detail_mapping, allowed, EXPECTED_DETAIL_HEADERS, "评分点行")
                )
                description = _field_text(row, detail_mapping, "description")
                command = _field_text(row, detail_mapping, "extra_description")
                requirement = _field_text(row, detail_mapping, "requirement")
                if not description:
                    errors.append(f"第 {row_number} 行：Aspect - Description 不能为空。")
                if not command:
                    errors.append(f"第 {row_number} 行：命令或操作说明不能为空。")
                if aspect_type == "M" and not requirement:
                    errors.append(f"第 {row_number} 行：M 类型评分点的 Requirement 不能为空。")
                max_mark, error = _decimal_from_cell(
                    _field_cell(row, detail_mapping, "max_mark"),
                    row_number=row_number,
                    field_label="Max Mark",
                    required=True,
                )
                if error:
                    errors.append(error)
                    max_mark = Decimal("0.00")

                judgement_options = []
                option_row_index = row_index + 1
                if aspect_type == "J":
                    for expected_score in (0, 1, 2, 3):
                        if option_row_index >= len(rows):
                            errors.append(f"第 {row_number} 行：J 类型评分点缺少 {expected_score} 分分档行。")
                            continue
                        option_row = rows[option_row_index]
                        option_row_number = option_row["row_number"]
                        if not _is_judgement_option_row(option_row, detail_mapping):
                            errors.append(f"第 {row_number} 行：J 类型评分点缺少 {expected_score} 分分档行。")
                            continue
                        errors.extend(
                            _row_unexpected_errors(
                                option_row,
                                detail_mapping,
                                {"judgement_score", "extra_description", "calculation_row", "max_mark"},
                                EXPECTED_DETAIL_HEADERS,
                                "J 分档行",
                            )
                        )
                        score, error = _decimal_from_cell(
                            _field_cell(option_row, detail_mapping, "judgement_score"),
                            row_number=option_row_number,
                            field_label="Judg Score",
                            required=True,
                        )
                        if error:
                            errors.append(error)
                            score = Decimal("-1.00")
                        if score != Decimal(expected_score).quantize(CENT):
                            errors.append(f"第 {option_row_number} 行：Judg Score 应为 {expected_score}。")
                        option_description = _field_text(option_row, detail_mapping, "extra_description")
                        if not option_description:
                            errors.append(f"第 {option_row_number} 行：J 分档说明不能为空。")
                        option_mark_cell = _field_cell(option_row, detail_mapping, "max_mark")
                        if option_mark_cell and option_mark_cell["normalized_text"]:
                            option_mark, error = _decimal_from_cell(
                                option_mark_cell,
                                row_number=option_row_number,
                                field_label="Max Mark",
                            )
                            if error:
                                errors.append(error)
                            else:
                                max_mark += option_mark
                        judgement_options.append(
                            ParsedJudgementOption(
                                score_value=score,
                                description=option_description,
                                source_row_number=option_row_number,
                                order=len(judgement_options),
                            )
                        )
                        option_row_index += 1
                aspect_number = aspect_count_by_subcriterion.get(current_subcriterion.code, 0) + 1
                aspect_count_by_subcriterion[current_subcriterion.code] = aspect_number
                current_subcriterion.aspects.append(
                    ParsedAspect(
                        code=f"{current_subcriterion.code}.{aspect_number}",
                        aspect_type=aspect_type,
                        description=description,
                        command=command,
                        requirement=requirement,
                        calculation_row=_field_text(row, detail_mapping, "calculation_row"),
                        max_mark=max_mark,
                        source_row_number=row_number,
                        order=aspect_order,
                        judgement_options=judgement_options,
                    )
                )
                aspect_order += 1
                row_index = option_row_index if aspect_type == "J" else row_index + 1
                continue

            errors.append(f"第 {row_number} 行：无法识别为子评分项、评分点或总分行。")
            row_index += 1

    if not subcriteria:
        errors.append("评分表至少需要一个子评分项。")
    if not any(subcriterion.aspects for subcriterion in subcriteria):
        errors.append("评分表至少需要一个评分点。")
    if not total_row_seen:
        errors.append("未找到 Total Marks: 总分行。")
        total_mark = Decimal("0.00")

    aspect_total = sum(
        (aspect.max_mark for subcriterion in subcriteria for aspect in subcriterion.aspects),
        Decimal("0.00"),
    ).quantize(CENT)
    if total_mark is not None and aspect_total != total_mark:
        errors.append(f"评分点分值总和为 {aspect_total}，与 Total Marks {total_mark} 不一致。")
    if total_mark is not None and module_mark and total_mark != module_mark:
        errors.append(f"Total Marks {total_mark} 与模块 Mark {module_mark} 不一致。")

    validation_report = {
        "errors": errors,
        "warnings": warnings,
        "checks": {
            "aspect_total": str(aspect_total),
            "total_mark": str(total_mark or Decimal("0.00")),
            "module_mark": str(module_mark),
        },
    }
    field_mapping = {
        "module": _mapping_payload(module_mapping, raw_snapshot, module_header_index, EXPECTED_MODULE_HEADERS)
        if module_header_index is not None
        else {},
        "details": _mapping_payload(detail_mapping, raw_snapshot, detail_header_index, EXPECTED_DETAIL_HEADERS)
        if detail_header_index is not None
        else {},
    }
    if errors:
        raise WorkbookParseError(errors)

    return ParsedWorkbook(
        module_code=module_code,
        module_name=module_name,
        module_mark=module_mark,
        title=title,
        total_mark=total_mark or Decimal("0.00"),
        subcriteria=subcriteria,
        raw_snapshot=raw_snapshot,
        field_mapping=field_mapping,
        validation_report=validation_report,
    )


def calculate_file_sha256(file_obj: BinaryIO) -> str:
    position = None
    try:
        position = file_obj.tell()
    except (AttributeError, OSError):
        position = None
    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass
    digest = hashlib.sha256()
    for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
        digest.update(chunk)
    if position is not None:
        try:
            file_obj.seek(position)
        except (AttributeError, OSError):
            pass
    return digest.hexdigest()


