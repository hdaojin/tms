from __future__ import annotations

from contextlib import contextmanager
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.core.files.storage import InMemoryStorage
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from archives.models import ArchiveAsset
from events.models import Event, EventModule
from knowledge.models import KnowledgeEvidence
from standards.models import SkillProject

from .forms import ScoringImportForm
from .models import (
    JudgementOption,
    ScoringAspect,
    ScoringParserConfig,
    ScoringParticipant,
    ScoringResult,
    ScoringScheme,
    ScoringSchemeImport,
)
from .parser import WorkbookParseError, parse_marking_workbook
from .services import confirm_scheme_import, default_parser_config, parse_scheme_upload, sync_parser_configs


@contextmanager
def archive_in_memory_storage():
    field = ArchiveAsset._meta.get_field("file")
    original_storage = field.storage
    field.storage = InMemoryStorage()
    try:
        yield
    finally:
        field.storage = original_storage


class ScoringImportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="coach")
        self.user.user_permissions.add(Permission.objects.get(codename="add_scoringscheme"))
        self.project = SkillProject.objects.create(code="NSM", name="网络系统管理")
        self.event = Event.objects.create(
            skill_project=self.project,
            event_type=Event.EventType.COMPETITION,
            name="全国选拔赛",
            code="NSM-SELECT",
            start_date=timezone.localdate(),
        )
        self.module = EventModule.objects.create(event=self.event, code="A", name="Linux")
        sync_parser_configs()
        self.parser_config = default_parser_config()

    def _cmp_workbook_upload(self, mutator=None, filename="cmp-marking.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Marking Scheme Import"
        sheet.cell(row=1, column=1).value = "IT Network Systems Administration"
        sheet.cell(row=3, column=1).value = "ID"
        sheet.cell(row=3, column=2).value = "Name"
        sheet.cell(row=3, column=11).value = "Mark"
        sheet.cell(row=4, column=1).value = "A"
        sheet.cell(row=4, column=2).value = "Linux environments"
        sheet.cell(row=4, column=11).value = 1.8
        headers = [
            "Sub\nCriterion\nID",
            "Sub Criterion\nName or Description",
            "Day of Marking",
            "Aspect\nType\nM = Meas\nJ = Judg",
            "Aspect - Description",
            "Judg Score",
            "Extra Aspect Description (Meas or Judg)\nOR\nJudgement Score Description (Judg only)",
            "Requirement\n(Measurement Only)",
            "WSOS Section",
            "Calculation Row \n(Export only)",
            "Max\nMark",
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=6, column=column).value = header
        rows = [
            (7, "A1", "fw1", "Day 1", None, None, None, None, None, None, None, None),
            (8, None, None, None, "M", "fw1 hostname\n is correct", None, "hostname -f", "fw1.example", "1.1", "C8", 0.2),
            (9, "A2", "mgmt1", "Day 1", None, None, None, None, None, None, None, None),
            (10, None, None, None, "M", "mgmt1 hostname is correct", None, "hostname -f", "mgmt1.example", "1.1", "C10", 0.2),
            (11, None, None, None, "J", "nftables Stateful Firewall Configuration", None, "nft list ruleset", "drop policy", "1.2", "C11", 1),
            (12, None, None, None, None, None, 0, "No valid output", None, None, None, None),
            (13, None, None, None, None, None, 1, "Basic ruleset exists", None, None, None, None),
            (14, None, None, None, None, None, 2, "Mostly meets requirements", None, None, None, 0.2),
            (15, None, None, None, None, None, 3, "Configuration is correct", None, None, None, 0.2),
            (16, None, None, None, None, None, None, None, None, None, "Total Marks:", 1.8),
        ]
        for row in rows:
            row_number = row[0]
            for column, value in enumerate(row[1:], start=1):
                sheet.cell(row=row_number, column=column).value = value
        if mutator:
            mutator(workbook, sheet)
        output = BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_cmp_template_workbook_parses_grouped_aspects_and_preserves_content(self):
        parsed = parse_marking_workbook(self._cmp_workbook_upload(), expected_module_code="A")
        aspects = [aspect for subcriterion in parsed.subcriteria for aspect in subcriterion.aspects]

        self.assertEqual(parsed.module_code, "A")
        self.assertEqual(parsed.module_name, "Linux environments")
        self.assertEqual(parsed.total_mark, Decimal("1.80"))
        self.assertEqual([subcriterion.code for subcriterion in parsed.subcriteria], ["A1", "A2"])
        self.assertEqual([aspect.code for aspect in aspects], ["A1.1", "A2.1", "A2.2"])
        self.assertEqual(aspects[0].description, "fw1 hostname\nis correct")
        self.assertEqual(aspects[0].calculation_row, "C8")
        self.assertEqual(aspects[-1].aspect_type, "J")
        self.assertEqual(aspects[-1].max_mark, Decimal("1.40"))
        self.assertEqual(
            [option.score_value for option in aspects[-1].judgement_options],
            [Decimal("0.00"), Decimal("1.00"), Decimal("2.00"), Decimal("3.00")],
        )
        self.assertEqual(parsed.validation_report["checks"]["aspect_total"], "1.80")
        self.assertIn("raw_snapshot", {"raw_snapshot": parsed.raw_snapshot})

    def test_parser_reports_all_blocking_errors(self):
        def break_template(_workbook, sheet):
            sheet.cell(row=9, column=1).value = "A3"
            sheet.cell(row=10, column=8).value = None
            sheet.cell(row=14, column=7).value = None
            sheet.cell(row=5, column=1).value = "unexpected block"
            sheet.cell(row=16, column=11).value = 9.9
            sheet.cell(row=18, column=1).value = "extra business content"

        with self.assertRaises(WorkbookParseError) as ctx:
            parse_marking_workbook(self._cmp_workbook_upload(break_template), expected_module_code="A")

        message = str(ctx.exception)
        self.assertIn("子评分项编号应为 A2", message)
        self.assertIn("Requirement 不能为空", message)
        self.assertIn("J 分档说明不能为空", message)
        self.assertIn("评分点分值总和", message)
        self.assertIn("模块定义与评分细则表头之间不能填写业务内容", message)
        self.assertIn("Total Marks 后不能再填写内容", message)
        self.assertGreaterEqual(len(ctx.exception.errors), 5)

    def test_parse_upload_creates_import_record_without_confirming_scheme(self):
        with archive_in_memory_storage():
            scheme_import = parse_scheme_upload(
                self.module,
                self._cmp_workbook_upload(),
                self.parser_config,
                user=self.user,
            )

        self.assertEqual(ScoringSchemeImport.objects.count(), 1)
        self.assertEqual(ArchiveAsset.objects.count(), 1)
        self.assertEqual(ScoringScheme.objects.count(), 0)
        self.assertEqual(ScoringAspect.objects.count(), 0)
        self.assertEqual(scheme_import.status, ScoringSchemeImport.Status.PARSED)
        self.assertEqual(scheme_import.parser_display_name, self.parser_config.display_name)
        self.assertEqual(scheme_import.parsed_payload["aspects_count"], 3)
        self.assertEqual(scheme_import.raw_snapshot["sheet_name"], "Marking Scheme Import")

    def test_parse_failure_does_not_create_asset_or_import_record(self):
        def remove_header(_workbook, sheet):
            sheet.cell(row=6, column=1).value = "Sub Criterion"

        with archive_in_memory_storage(), self.assertRaises(WorkbookParseError):
            parse_scheme_upload(
                self.module,
                self._cmp_workbook_upload(remove_header),
                self.parser_config,
                user=self.user,
            )

        self.assertEqual(ArchiveAsset.objects.count(), 0)
        self.assertEqual(ScoringSchemeImport.objects.count(), 0)
        self.assertEqual(ScoringScheme.objects.count(), 0)

    def test_confirm_import_creates_scheme_aspects_judgement_options_and_evidence(self):
        with archive_in_memory_storage():
            scheme_import = parse_scheme_upload(
                self.module,
                self._cmp_workbook_upload(),
                self.parser_config,
                user=self.user,
            )
            scheme = confirm_scheme_import(scheme_import, user=self.user)

        self.assertEqual(scheme.total_mark, Decimal("1.80"))
        self.assertEqual(ScoringAspect.objects.filter(scheme=scheme).count(), 3)
        self.assertEqual(JudgementOption.objects.filter(aspect__scheme=scheme).count(), 4)
        self.assertEqual(scheme.aspects.get(code="A1.1").calculation_row, "C8")
        self.assertEqual(scheme_import.status, ScoringSchemeImport.Status.CONFIRMED)
        self.assertEqual(scheme_import.scheme, scheme)
        self.assertEqual(
            KnowledgeEvidence.objects.filter(
                skill_project=self.project,
                source_type=KnowledgeEvidence.SourceType.SCORING_ASPECT,
                review_status=KnowledgeEvidence.ReviewStatus.APPROVED,
            ).count(),
            3,
        )

    def test_confirm_import_overwrites_existing_scheme_only_without_results(self):
        with archive_in_memory_storage():
            first_import = parse_scheme_upload(self.module, self._cmp_workbook_upload(), self.parser_config, user=self.user)
            scheme = confirm_scheme_import(first_import, user=self.user)
            second_import = parse_scheme_upload(
                self.module,
                self._cmp_workbook_upload(filename="cmp-marking-2.xlsx"),
                self.parser_config,
                user=self.user,
            )
            overwritten = confirm_scheme_import(second_import, user=self.user)

        self.assertEqual(overwritten.pk, scheme.pk)
        self.assertEqual(ScoringScheme.objects.count(), 1)
        self.assertEqual(ScoringAspect.objects.filter(scheme=scheme).count(), 3)

    def test_confirm_import_blocks_overwrite_when_results_exist(self):
        with archive_in_memory_storage():
            first_import = parse_scheme_upload(self.module, self._cmp_workbook_upload(), self.parser_config, user=self.user)
            scheme = confirm_scheme_import(first_import, user=self.user)
            participant = ScoringParticipant.objects.create(
                scheme=scheme,
                external_identifier="competitor-1",
                display_name="Competitor 1",
            )
            ScoringResult.objects.create(participant=participant, aspect=scheme.aspects.first(), score_awarded=Decimal("0.10"))
            second_import = parse_scheme_upload(
                self.module,
                self._cmp_workbook_upload(filename="cmp-marking-2.xlsx"),
                self.parser_config,
                user=self.user,
            )

        with self.assertRaises(ValidationError):
            confirm_scheme_import(second_import, user=self.user)

    def test_parser_config_sync_and_import_form_default(self):
        config = ScoringParserConfig.objects.get(parser_key="cmp_single_module_v1")
        config.display_name = "自定义 CMP 名称"
        config.alias = "CMP-STRICT"
        config.save()

        sync_parser_configs()
        config.refresh_from_db()
        form = ScoringImportForm()

        self.assertEqual(config.display_name, "自定义 CMP 名称")
        self.assertEqual(form.fields["parser_config"].initial, config.pk)
        self.assertIn(config, list(form.fields["parser_config"].queryset))
        self.assertEqual(form.fields["parser_config"].label_from_instance(config), "自定义 CMP 名称")

    def test_import_form_orders_event_modules_and_uses_event_name_labels(self):
        old_event = Event.objects.create(
            skill_project=self.project,
            event_type=Event.EventType.COMPETITION,
            name="旧赛事",
            code="OLD-EVENT",
            start_date=timezone.localdate() - timedelta(days=10),
        )
        EventModule.objects.create(event=old_event, code="Z", name="Old Module", order=1)
        latest_event = Event.objects.create(
            skill_project=self.project,
            event_type=Event.EventType.COMPETITION,
            name="最新赛事",
            code="LATEST-EVENT",
            start_date=timezone.localdate() + timedelta(days=1),
        )
        latest_second = EventModule.objects.create(event=latest_event, code="B", name="Second", order=2)
        latest_first = EventModule.objects.create(event=latest_event, code="C", name="First", order=1)

        form = ScoringImportForm()
        event_modules = list(form.fields["event_module"].queryset)

        self.assertEqual(event_modules[:2], [latest_first, latest_second])
        self.assertEqual(form.fields["event_module"].initial, latest_first.pk)
        self.assertEqual(form.fields["event_module"].label_from_instance(latest_first), "最新赛事 / C - First")
        self.assertNotIn("LATEST-EVENT", form.fields["event_module"].label_from_instance(latest_first))

    def test_import_view_parses_to_preview_and_template_download_works(self):
        self.client.force_login(self.user)
        with archive_in_memory_storage():
            response = self.client.post(
                reverse("scoring:scheme_import"),
                {
                    "event_module": self.module.pk,
                    "parser_config": self.parser_config.pk,
                    "file": self._cmp_workbook_upload(),
                },
            )

        self.assertEqual(response.status_code, 302)
        scheme_import = ScoringSchemeImport.objects.get()
        self.assertEqual(response.url, reverse("scoring:scheme_import_preview", args=[scheme_import.pk]))
        self.assertEqual(ScoringScheme.objects.count(), 0)

        preview = self.client.get(response.url)
        self.assertContains(preview, "确认生成评分方案")
        self.assertContains(preview, "nftables Stateful Firewall Configuration")

        template_response = self.client.get(reverse("scoring:parser_template", args=[self.parser_config.parser_key]))
        self.assertEqual(template_response.status_code, 200)
        self.assertEqual(
            template_response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )




